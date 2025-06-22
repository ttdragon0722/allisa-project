from openpyxl import load_workbook
from abc import  ABC, abstractmethod
import re
from .debug import Debug

class Excel(ABC):
    def __init__(self):
        self.path:str = None
        
    def get_fileName(self):
        return self.path.split("/")[-1]
    
    @abstractmethod
    def import_excel(self,path: str):
        try:
            self.workbook = load_workbook(path)
        except FileNotFoundError:
            raise FileNotFoundError(f"找不到檔案: {path}")
        except Exception as e:
            raise Exception(f"讀取 Excel 檔案失敗: {e}")
    
    @abstractmethod
    def get_table(self):
        pass
    
    @abstractmethod
    def get_pieces(self):
        pass
    
    @abstractmethod
    def get_keywords(self):
        pass
    
    

class SOMReader(Excel):
    def __init__(self):
        super().__init__() 
    
    def import_excel(self, path):
        super().import_excel(path)
    
        if "工程專用" not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '工程專用' 不存在，請確認名稱是否正確。可用工作表: {self.workbook.sheetnames}")

        self.worksheet = self.workbook["工程專用"]
        self.path = path
    
    
    def get_table(self):
        """回傳整張表(含標題)的二維清單"""
        data = []
        for row in self.worksheet.iter_rows(values_only=True):
            data.append(list(row))
        return data


    def get_pieces(self):
        """去除標題，只取 A、B 欄，攤平成一維清單（不拆逗號）"""
        data = self.get_table()
        if len(data) <= 1:
            return []

        pieces = []
        # 只取 A、B 欄（第0,1欄），去掉標題(第一列)
        for row in data[1:]:
            for col_index in [0, 1]:
                if col_index < len(row):
                    val = row[col_index]
                    if val is not None and val != "":
                        pieces.append(str(val).strip())
        return pieces


    def get_keywords(self):
        """拆逗號，去除括號內容，去重後回傳關鍵字"""
        pieces = self.get_pieces()
        seen = set()
        result = []

        for item in pieces:
            parts = item.split(',')
            for part in parts:
                cleaned = re.sub(r"\(.*?\)", "", part).strip()  # 去除括號
                if cleaned and cleaned not in seen:
                    seen.add(cleaned)
                    result.append(cleaned)

        return result
        

class ExcelReader:
    def __init__(self):
        self.path:str = None
        
    def get_fileName(self):
        return self.path.split("/")[-1]
    
    def import_excel(self,path):
        try:
            self.workbook = load_workbook(path)
        except FileNotFoundError:
            raise FileNotFoundError(f"找不到檔案: {path}")
        except Exception as e:
            raise Exception(f"讀取 Excel 檔案失敗: {e}")

        if "零件位置" not in self.workbook.sheetnames:
            raise ValueError(f"工作表 '零件位置' 不存在，請確認名稱是否正確。可用工作表: {self.workbook.sheetnames}")

        self.worksheet = self.workbook["零件位置"]
        self.path = path
        
    def get_table(self):
        ws = self.worksheet
        boundary_rows = []

        # 找出高度小於 8 的列作為邊界
        for row in range(1, ws.max_row + 1):
            row_dim = ws.row_dimensions.get(row)
            height = row_dim.height if row_dim and row_dim.height is not None else None
            if height is not None and height < 8:
                boundary_rows.append(row)

        # 如果邊界不足，回傳空表
        if len(boundary_rows) < 2:
            return []

        table_data = []

        # 擷取邊界之間的資料（不含邊界行），從 C 到 F 欄
        for i in range(len(boundary_rows) - 1):
            start = boundary_rows[i] + 1 #這裡要標題
            end = boundary_rows[i + 1] - 1
            if start <= end:
                for row in range(start, end + 1):
                    row_data = []
                    for col in range(3, 7):  # C=3, D=4, E=5, F=6
                        val = ws.cell(row=row, column=col).value
                        row_data.append(val)
                    table_data.append(row_data)

        return table_data

    @Debug.event("get_pieces","cyan")
    def get_pieces(self):
        ws = self.worksheet
        boundary_rows = []

        # 找出高度小於 5 的列（作為邊界）
        for row in range(1, ws.max_row + 1):
            row_dim = ws.row_dimensions.get(row)
            height = row_dim.height if row_dim and row_dim.height is not None else None
            if height is not None and height < 8:
                boundary_rows.append(row)

        # 如果邊界數量小於 2，沒有中間資料可取
        if len(boundary_rows) < 2:
            return []

        result = []

        # 擷取「每對相鄰邊界之間」的資料（不含邊界行）
        for i in range(len(boundary_rows) - 1):
            start = boundary_rows[i] + 2  # 去除header 從+2行開始讀
            end = boundary_rows[i + 1] - 1
            if start <= end:
                for row in range(start, end + 1):
                    c_val = ws.cell(row=row, column=3).value  # C欄
                    d_val = ws.cell(row=row, column=4).value  # D欄
                    if c_val is not None:
                        result.append(c_val)
                    if d_val is not None:
                        result.append(d_val)

        print(result)
        return result
    
    def get_keywords(self):
        pieces = self.get_pieces()
        seen = set()
        result = []

        for item in pieces:
            if item:
                parts = item.split(',')
                for part in parts:
                    cleaned = part.strip()

                    # ✅ 去除括號內容，例如 JP1(1-2) -> JP1（不轉小寫）
                    cleaned = re.sub(r"\(.*?\)", "", cleaned).strip()

                    if cleaned and cleaned not in seen:
                        seen.add(cleaned)
                        result.append(cleaned)

        return result
